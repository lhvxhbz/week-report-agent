"""File scanning and reading module for weekly report agent.

Provides scan_folder() to discover recently modified files and read_file()
to extract content from various file formats.
"""

import hashlib
import os
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Well-known directories to skip
_IGNORED_DIRS = frozenset({
    "node_modules",
    "__pycache__",
    "venv",
    ".venv",
    "env",
    ".env",
    "dist",
    "build",
    ".git",
    ".svn",
    ".hg",
    ".idea",
    ".vscode",
    ".vs",
    ".mypy_cache",
    ".pytest_cache",
    ".tox",
    ".eggs",
    "egg-info",
    ".cache",
    ".next",
    ".nuxt",
    "coverage",
    ".sass-cache",
    "bower_components",
    "jspm_packages",
    ".gradle",
    "target",
    "bin",
    "obj",
})

# Maximum file size: 1 MB
MAX_FILE_SIZE = 1 * 1024 * 1024

# Maximum recursion depth
MAX_DEPTH = 5

# Supported text file extensions for read_file()
_TEXT_EXTENSIONS = frozenset({
    ".txt",
    ".md",
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".html",
    ".htm",
    ".css",
    ".json",
    ".csv",
    ".yaml",
    ".yml",
    ".toml",
    ".ini",
    ".cfg",
    ".conf",
    ".sh",
    ".bat",
    ".ps1",
    ".xml",
    ".svg",
    ".sql",
    ".r",
    ".rb",
    ".go",
    ".rs",
    ".java",
    ".kt",
    ".swift",
    ".c",
    ".cpp",
    ".h",
    ".hpp",
    ".cs",
    ".lua",
    ".php",
    ".pl",
    ".log",
    ".env",
    ".gitignore",
    ".dockerignore",
    ".editorconfig",
    ".prettierrc",
    ".eslintrc",
    ".lock",
    ".makefile",
    ".cmake",
})

# Docx extension
_DOCX_EXTENSIONS = frozenset({".docx"})

# PDF extension
_PDF_EXTENSIONS = frozenset({".pdf"})

# Excel extensions
_EXCEL_EXTENSIONS = frozenset({".xlsx", ".xls"})

# Encoding order: most common first for faster detection
_ENCODINGS = ("utf-8", "utf-8-sig", "gbk", "gb2312", "latin-1")

# Thread-safe encoding cache: maps file_path -> working encoding name
_ENCODING_CACHE: Dict[str, str] = {}
_ENCODING_CACHE_LOCK = threading.Lock()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _should_skip_dir(dirname: str) -> bool:
    """Return True if directory should be skipped."""
    # Skip hidden directories
    if dirname.startswith("."):
        return True
    # Skip well-known ignored directories
    if dirname in _IGNORED_DIRS:
        return True
    # Also match patterns like package.egg-info
    if ".egg-info" in dirname:
        return True
    return False


def _format_mtime(timestamp: float) -> str:
    """Format a file's mtime as 'YYYY-MM-DD HH:MM'."""
    return datetime.fromtimestamp(timestamp).strftime("%Y-%m-%d %H:%M")


def _read_docx(file_path: Path) -> str:
    """Read a .docx file using python-docx. Returns plain text content."""
    try:
        from docx import Document  # type: ignore[import-untyped]

        doc = Document(str(file_path))
        paragraphs = [p.text for p in doc.paragraphs]
        return "\n".join(paragraphs)
    except ImportError:
        return "[Error] python-docx is not installed. Run: pip install python-docx"
    except Exception as exc:
        return f"[Error] Failed to read docx file: {exc}"


def _read_pdf(file_path: Path) -> str:
    """Read a .pdf file using PyPDF2. Returns plain text content."""
    try:
        from PyPDF2 import PdfReader

        reader = PdfReader(str(file_path))
        text_parts = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                text_parts.append(text)
        return "\n\n".join(text_parts)
    except ImportError:
        return "[Error] PyPDF2 is not installed. Run: pip install PyPDF2"
    except Exception as exc:
        return f"[Error] Failed to read PDF file: {exc}"


def _read_excel(file_path: Path) -> str:
    """Read an .xlsx/.xls file using openpyxl. Returns text representation."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(file_path), read_only=True, data_only=True)
        text_parts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text_parts.append(f"## Sheet: {sheet_name}")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join(
                    str(cell) if cell is not None else "" for cell in row
                )
                if row_text.strip():
                    text_parts.append(row_text)
        wb.close()
        return "\n".join(text_parts)
    except ImportError:
        return "[Error] openpyxl is not installed. Run: pip install openpyxl"
    except Exception as exc:
        return f"[Error] Failed to read Excel file: {exc}"


def get_file_hash(file_path: str) -> str:
    """根据文件路径和修改时间生成哈希值，用于变更检测。

    Parameters
    ----------
    file_path : str
        文件路径。

    Returns
    -------
    str
        文件的哈希值（MD5）。
    """
    try:
        stat = os.stat(file_path)
        # 使用路径 + 修改时间 + 文件大小作为缓存键
        key = f"{file_path}:{stat.st_mtime}:{stat.st_size}"
        return hashlib.md5(key.encode()).hexdigest()
    except OSError:
        return hashlib.md5(file_path.encode()).hexdigest()


def scan_folder(folder_path: str, days: int = 7) -> List[Dict]:
    """Scan a folder and return files modified within the last *days* calendar days.

    "days" means calendar days, not 24-hour periods.
    For example, if today is May 28 and days=1, it scans files from May 27 00:00 to May 27 23:59.
    If days=7, it scans files from May 21 00:00 to May 27 23:59.

    Parameters
    ----------
    folder_path : str
        Absolute or relative path to the folder to scan.
    days : int
        Number of calendar days to look back (default 7).

    Returns
    -------
    list[dict]
        Each dict contains: path, name, relative, modified, size, ext.
    """
    root = Path(folder_path).resolve()
    if not root.is_dir():
        return []

    # 计算日历天数的截止时间：days天前的0点
    # 例如：今天5月28日，days=1 -> 5月27日00:00:00
    # 例如：今天5月28日，days=7 -> 5月21日00:00:00
    now = datetime.now()
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    cutoff = today_start - timedelta(days=days - 1)
    cutoff_ts = cutoff.timestamp()
    results: List[Dict] = []

    def _walk(current: Path, depth: int) -> None:
        if depth > MAX_DEPTH:
            return
        try:
            entries = sorted(
                os.scandir(str(current)), key=lambda e: e.name
            )
        except PermissionError:
            return
        except OSError:
            return

        for entry in entries:
            try:
                # Directory handling - use DirEntry.is_dir() (no extra syscall)
                if entry.is_dir(follow_symlinks=False):
                    if _should_skip_dir(entry.name):
                        continue
                    _walk(Path(entry.path), depth + 1)
                    continue

                # File handling - use DirEntry.is_file() (no extra syscall)
                if not entry.is_file(follow_symlinks=False):
                    continue

                # Get file stats - DirEntry.stat() uses cached info on Windows
                try:
                    stat = entry.stat(follow_symlinks=False)
                    mtime = stat.st_mtime
                    size = stat.st_size
                except OSError:
                    continue

                # Skip files older than cutoff
                if mtime < cutoff_ts:
                    continue

                # Skip files exceeding max size
                if size > MAX_FILE_SIZE:
                    continue

                # Build result
                entry_path = Path(entry.path)
                ext = entry_path.suffix.lower()
                rel = entry_path.relative_to(root)
                results.append(
                    {
                        "path": entry.path,
                        "name": entry.name,
                        "relative": str(rel),
                        "modified": _format_mtime(mtime),
                        "size": size,
                        "ext": ext,
                    }
                )
            except PermissionError:
                continue
            except OSError:
                continue

    _walk(root, depth=0)
    return results


def scan_folder_all(folder_path: str) -> List[Dict]:
    """Scan a folder and return ALL files (no time filter).

    Used for displaying the complete file list in the UI.

    Parameters
    ----------
    folder_path : str
        Absolute or relative path to the folder to scan.

    Returns
    -------
    list[dict]
        Each dict contains: path, name, relative, modified, size, ext.
    """
    root = Path(folder_path).resolve()
    if not root.is_dir():
        return []

    results: List[Dict] = []

    def _walk(current: Path, depth: int) -> None:
        if depth > MAX_DEPTH:
            return
        try:
            entries = sorted(
                os.scandir(str(current)), key=lambda e: e.name
            )
        except PermissionError:
            return
        except OSError:
            return

        for entry in entries:
            try:
                # Directory handling
                if entry.is_dir(follow_symlinks=False):
                    if _should_skip_dir(entry.name):
                        continue
                    _walk(Path(entry.path), depth + 1)
                    continue

                # File handling
                if not entry.is_file(follow_symlinks=False):
                    continue

                # Get file stats
                try:
                    stat = entry.stat(follow_symlinks=False)
                    mtime = stat.st_mtime
                    size = stat.st_size
                except OSError:
                    continue

                # Skip files exceeding max size
                if size > MAX_FILE_SIZE:
                    continue

                # Build result
                entry_path = Path(entry.path)
                ext = entry_path.suffix.lower()
                rel = entry_path.relative_to(root)
                results.append(
                    {
                        "path": entry.path,
                        "name": entry.name,
                        "relative": str(rel),
                        "modified": _format_mtime(mtime),
                        "size": size,
                        "ext": ext,
                    }
                )
            except PermissionError:
                continue
            except OSError:
                continue

    _walk(root, depth=0)
    return results


def read_file(file_path: str) -> str:
    """Read the content of a file, supporting multiple formats.

    Supported formats include .txt, .md, .py, .js, .ts, .tsx, .html, .css,
    .json, .csv, .docx, and many more text-based formats.

    Parameters
    ----------
    file_path : str
        Path to the file to read.

    Returns
    -------
    str
        File content as a string, or an error/message string on failure.
    """
    path = Path(file_path)

    # Validate existence
    if not path.exists():
        return f"[Error] File not found: {file_path}"

    if not path.is_file():
        return f"[Error] Not a file: {file_path}"

    # Check file size before reading
    try:
        size = path.stat().st_size
    except OSError as exc:
        return f"[Error] Cannot access file: {exc}"

    if size == 0:
        return ""

    if size > MAX_FILE_SIZE:
        return (
            f"[Skipped] File exceeds 1 MB limit "
            f"({size / 1024 / 1024:.2f} MB): {file_path}"
        )

    ext = path.suffix.lower()

    # Handle .docx
    if ext in _DOCX_EXTENSIONS:
        return _read_docx(path)

    # Handle .pdf
    if ext in _PDF_EXTENSIONS:
        return _read_pdf(path)

    # Handle .xlsx / .xls
    if ext in _EXCEL_EXTENSIONS:
        return _read_excel(path)

    # Handle text-based files
    if ext in _TEXT_EXTENSIONS or ext == "":
        return _read_text_file(path)

    return f"[Skipped] Unsupported file format: {ext}"


def read_files_concurrent(
    file_paths: List[str],
    max_workers: int = 4,
) -> Dict[str, str]:
    """Read multiple files concurrently using a thread pool.

    Parameters
    ----------
    file_paths : list[str]
        List of file paths to read.
    max_workers : int
        Maximum number of concurrent reader threads (default 4).

    Returns
    -------
    dict[str, str]
        Mapping of file_path -> content. Error strings are included as-is.
    """
    if not file_paths:
        return {}

    results: Dict[str, str] = {}

    # Cap workers to number of files
    effective_workers = min(max_workers, len(file_paths))

    with ThreadPoolExecutor(max_workers=effective_workers) as executor:
        future_to_path = {
            executor.submit(read_file, fp): fp for fp in file_paths
        }
        for future in as_completed(future_to_path):
            fp = future_to_path[future]
            try:
                results[fp] = future.result()
            except Exception as exc:
                results[fp] = f"[Error] Unexpected error reading {fp}: {exc}"

    return results


def _read_text_file(path: Path) -> str:
    """Read a text file with encoding fallback and caching."""
    path_str = str(path)

    # Check cache first
    with _ENCODING_CACHE_LOCK:
        cached_enc = _ENCODING_CACHE.get(path_str)

    if cached_enc is not None:
        try:
            return path.read_text(encoding=cached_enc)
        except (UnicodeDecodeError, OSError):
            # Cache miss - encoding no longer valid, fall through
            pass

    # Try each encoding
    for enc in _ENCODINGS:
        try:
            content = path.read_text(encoding=enc)
            # Cache the working encoding
            with _ENCODING_CACHE_LOCK:
                _ENCODING_CACHE[path_str] = enc
            return content
        except UnicodeDecodeError:
            continue
        except OSError as exc:
            return f"[Error] Failed to read file: {exc}"

    return f"[Error] Unable to decode file with supported encodings: {path}"
